import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

EXCEL_URL = './archivos-excel/Tarifario-Gral-Septiembre-mato.xlsx'

HOJAS_Y_COLUMNAS = {

    'Edding': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'Fenicio-Luminatec': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'LOBO ESTA': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'Mak Nutrition': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'VacaValiente': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'Gral-Muchas marcas': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'Gral-Muchas marcas 2': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto'
    ],

    'Juice market': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD',
        'Etiquetado',
        'Materiales'
    ],

    'Cascanueces': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD',
        'Etiquetado',
    ],

    'Craft Moments': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'Armado x Bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'Tarifario Rabieta MKT': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',
    ],

    'Zulki': [
        'ENVIO',
        'Excedente Kg',
        'Excedente x bulto',
        'PICK&PAD x bulto',
        'Etiquetado x bulto',
        'Materiales x bulto',        
    ]
}

# Costo x CEDOL
SERVICIOS_A_MODIFICAR = [
    'NORMAL AMBA/NEXT DAY',
    'Mercado_envios (despacho)',
    'Retira por desposito',
    'Normal Amba/Next Day',
    'Despacho',
    'Especial',
    'Pick Up'
]

# Costo x Flex
SERVICIO_FLEX = [
    'Same day/Flex'
]



def process_etl():

    wb = load_workbook(EXCEL_URL)
    
    for sheet, columnas_a_modificar in HOJAS_Y_COLUMNAS.items():

        
        ws = wb[sheet]
        # Leer los datos de un DataFrame de pandas
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0].str.strip()  # Eliminar espacios en blanco de los nombres de columnas
        df = df[1:]
        
        print(f"Columnas en la hoja '{sheet}':", df.columns)  # Imprimir columnas para verificar


        # Limpiar y convertir campos a decimal, solo si la columna existe
        for columna in columnas_a_modificar:
            if columna in df.columns:  # Verifica si la columna está en el DataFrame
                df[columna] = df[columna].astype(str).str.replace('$', '').str.replace(' ', '')
                df[columna] = pd.to_numeric(df[columna], errors='coerce')
            else:
                print(f"Columna '{columna}' no encontrada en la hoja '{sheet}', omitiendo.")

        # Descuento "cedol"
        mask = df['Tipo de servicio'].isin(SERVICIOS_A_MODIFICAR)

        # Descuento "flex"
        mask2 = df['Tipo de servicio'].isin(SERVICIO_FLEX)

        # Aplicar aumento solo si la columna 'Tipo de servicio' existe
        if 'Tipo de servicio' in df.columns:
            for columna in columnas_a_modificar:
                if columna in df.columns:  # Verifica si la columna está en el DataFrame
                    df.loc[mask, columna] = df.loc[mask, columna] * 1.10
                    df.loc[mask2, columna] = df.loc[mask2, columna] * 1.20

            # Aplicar formato solo a las filas que cumplen con `mask`
            for columna in columnas_a_modificar:
                if columna in df.columns:  # Verifica si la columna está en el DataFrame
                    df.loc[mask, columna] = df.loc[mask, columna].apply(lambda x: f'$ {x:.2f}' if pd.notnull(x) else x)
                    df.loc[mask2, columna] = df.loc[mask2, columna].apply(lambda x: f'$ {x:.2f}' if pd.notnull(x) else x)
                    
        else:
            print(f"Columna 'Tipo de servicio' no encontrada en la hoja '{sheet}', omitiendo modificaciones.")


        # Desfusionar celdas antes de escribir nuevos valores
        merged_cells = list(ws.merged_cells.ranges)
        for merged_cell in merged_cells:
            ws.unmerge_cells(str(merged_cell))


        # Escribir los datos actualizados en la hoja
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    


    wb.save('precios_actualizados.xlsx')

process_etl()